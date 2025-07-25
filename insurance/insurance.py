import asyncio
import openpyxl
from playwright.async_api import async_playwright

EXCEL_FILE = "ICICI Prudential Life.xlsx"  # Make sure this path is correct

async def main():
    # Load workbook and sheet
    wb = openpyxl.load_workbook(EXCEL_FILE)
    sheet = wb.active  # Or sheet by name: wb["Sheet1"]

    header_row_index = 1
    headers = [str(cell.value).strip() for cell in sheet[header_row_index]]
    #print("Cleaned headers:", headers)
    name_col = headers.index("Product")
    print(f"Product column index: {name_col}")
    uin_col = headers.index("UIN")
    # Check if 'URL' column exists
    if "URL" not in headers:
        headers.append("URL")
        sheet.cell(row=header_row_index, column=len(headers)).value = "URL"
        print("Added 'URL' column to headers.",headers)
    url_col_index = headers.index("URL")  # ✅ FIX: column index

    # Start from the row after the header
    for row in sheet.iter_rows(min_row=header_row_index + 1):  # ✅ read full rows (cells)
        product_name = row[name_col].value
        product_uin = row[uin_col].value

        if not product_name or not product_uin:
            continue

        search_query = f"{product_name} (UIN:{product_uin}) download brochure"
        print(search_query)

        async with async_playwright() as p:
            browser = await p.chromium.launch(headless=False)
            page = await browser.new_page()

            search_url = f"https://www.bing.com/search?q={search_query}"
            await page.goto(search_url, wait_until="load")
            await page.wait_for_timeout(3000)
            #await page.screenshot(path=f"screenshots/{product_uin}_search.png", full_page=True)

            await page.wait_for_selector("li.b_algo h2 a")
            top_links = await page.locator("li.b_algo h2 a").all()
            first_link = top_links[0] if top_links else None

            row_num = row[0].row  # ✅ correct dynamic row number

            if first_link:
                href = await first_link.get_attribute("href")
                title = await first_link.inner_text()
                print(f"🔗 Top Link: {title} — {href}")
                sheet.cell(row=row_num, column=url_col_index + 1).value = href
            else:
                print("❌ No search result links found.")
                sheet.cell(row=row_num, column=url_col_index + 1).value = "No link found"
            # Save the workbook after each row
            wb.save(EXCEL_FILE)
            await browser.close()

        



asyncio.run(main())
